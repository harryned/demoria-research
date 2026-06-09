#!/usr/bin/env python3
"""
Demoria Research — populate Birth Tracker tiles from the manual input sheet.

Reads Birth_Tracker_Input.xlsx and writes births into public/births_data.json:
  • Monthly: takes the leading run of months filled in BOTH 2025 and 2026,
    sums each, sets b25/b26/mon  -> tile shows "first N months, ±X%".
  • Annual only: if no monthly 2026 but FY2025+FY2026 present, sets the
    annual block (ba) -> tile shows "2026 full year, ±X%".
  • Blank rows are left as-is ("awaiting release").

Usage:
  python3 _births_from_sheet.py            # dry run (prints what would change)
  python3 _births_from_sheet.py --write    # write into births_data.json
"""
import json, sys
from openpyxl import load_workbook

SHEET='Birth_Tracker_Input_EXPANDED.xlsx'; DATA='public/births_data.json'
M24=range(4,16)    # cols D..O   (2024 Jan..Dec)
M25=range(16,28)   # cols P..AA  (2025 Jan..Dec)
M26=range(28,40)   # cols AB..AM (2026 Jan..Dec)
FY24=40; FY25=41; FY26=42   # cols AN, AO, AP
SRC=43             # col AQ
FORCE_ANNUAL={'MDA'}   # countries whose monthly data is too sparse — use annual FY2024 vs FY2025

def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return int(round(v))
    s=str(v).strip().replace(',','').replace(' ','')
    if not s: return None
    try: return int(round(float(s)))
    except: return None

def parse():
    wb=load_workbook(SHEET, data_only=True)
    ws=wb['Births']
    out={}
    for row in ws.iter_rows(min_row=3):
        iso=row[0].value
        if not iso or len(str(iso).strip())!=3: continue   # skip region banners / blanks
        iso=str(iso).strip()
        m24=[num(row[c-1].value) for c in M24]
        m25=[num(row[c-1].value) for c in M25]
        m26=[num(row[c-1].value) for c in M26]
        src=row[SRC-1].value
        # Forced-annual reporters: monthly data is too sparse to trust, so use the
        # latest two FULL annual years (FY2024 vs FY2025) regardless of any monthly cells.
        if iso in FORCE_ANNUAL:
            a,b=num(row[FY24-1].value),num(row[FY25-1].value)
            if a and b and a>0 and b>0:
                out[iso]={'mode':'annual','prev':a,'cur':b,'year':2025,'src':src}
            continue
        # leading run of months present in BOTH years
        run=0
        for a,b in zip(m25,m26):
            if a is not None and b is not None: run+=1
            else: break
        if run>0:
            out[iso]={'mode':'monthly','b25':sum(m25[:run]),'b26':sum(m26[:run]),'mon':run,'src':src}
            continue
        # Complete-year fallback: full 2025 (all 12 months) but no 2026 yet -> show 2025 vs 2024.
        # Only fires on COMPLETE years, so no partial-vs-full distortion. Auto-upgrades to the
        # 2026-vs-2025 monthly view as soon as 2026 months are added.
        n25=sum(1 for x in m25 if x is not None)
        n26=sum(1 for x in m26 if x is not None)
        if n25==12 and n26==0:
            n24=sum(1 for x in m24 if x is not None)
            fy24=num(row[FY24-1].value)
            total24=fy24 if (fy24 and fy24>0) else (sum(m24) if n24==12 else None)
            if total24 and total24>0:
                out[iso]={'mode':'annual','prev':total24,'cur':sum(m25),'year':2025,'src':src}
                continue
        # Annual fallback — ONLY for pure-annual reporters (no monthly cells anywhere, e.g. China).
        # Monthly countries' FY columns are auto-sums that may be partial-year, so never compare them.
        if any(x is not None for x in m24+m25+m26): continue
        # Use the latest year Y where both FY_Y and FY_{Y-1} are present.
        # (China only publishes annual, one year behind — its latest is 2025 vs 2024, not 2026.)
        fy={2024:num(row[FY24-1].value),2025:num(row[FY25-1].value),2026:num(row[FY26-1].value)}
        for y in (2026,2025):
            if fy[y] and fy[y-1] and fy[y]>0 and fy[y-1]>0:
                out[iso]={'mode':'annual','prev':fy[y-1],'cur':fy[y],'year':y,'src':src}
                break
    return out

def merge(parsed, write=False):
    data=json.load(open(DATA,encoding='utf-8'))
    by={c['iso']:c for c in data['countries']}
    changed=[]
    for iso,r in parsed.items():
        c=by.get(iso)
        if not c: print(f"  ! {iso} not in dataset, skipped"); continue
        if r['mode']=='monthly':
            c['b25']=r['b25']; c['b26']=r['b26']; c['mon']=r['mon']; c['ba']=None
            chg=(r['b26']-r['b25'])/r['b25']*100 if r['b25'] else 0
            changed.append((iso,f"{r['b26']:,} vs {r['b25']:,} ({r['mon']}mo)  {chg:+.1f}%"))
        else:
            c['ba']={'year':r['year'],'prev':r['prev'],'cur':r['cur']}; c['b25']=None; c['b26']=None; c['mon']=None
            chg=(r['cur']-r['prev'])/r['prev']*100 if r['prev'] else 0
            changed.append((iso,f"FY {r['cur']:,} vs {r['prev']:,}  {chg:+.1f}%"))
        if r.get('src'): c['births_source']=str(r['src'])
    if write and changed:
        data['updated']=__import__('datetime').date.today().isoformat()
        json.dump(data,open(DATA,'w',encoding='utf-8'),ensure_ascii=False,separators=(',',':'))
    return changed

if __name__=='__main__':
    write='--write' in sys.argv
    parsed=parse()
    changed=merge(parsed, write=write)
    if not changed:
        print("No births entered in the sheet yet — nothing to update.")
    else:
        print(f"{'WROTE' if write else 'Would update'} {len(changed)} countries:")
        for iso,msg in changed: print(f"  {iso}  {msg}")
        if not write: print("\nRun again with --write to apply, then deploy (cp + git push).")
