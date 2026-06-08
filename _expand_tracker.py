#!/usr/bin/env python3
"""Expand Birth_Tracker_Input.xlsx to all 236 DHI countries/territories and fill
European nations' monthly + annual births from Eurostat demo_fmonth.

- Preserves every existing cell value AND style exactly (rebuild, not insert).
- Adds missing countries into the correct section, alphabetical by name.
- Fills Eurostat 2024/2025 monthly + FY totals + provenance for European rows.
- Never overwrites a non-empty existing cell.
"""
import json, openpyxl, shutil, re
from copy import copy
from openpyxl.utils import get_column_letter

def remap_formula(v, oldr, newr):
    """Rewrite self-row cell references (e.g. P32, AN23) from oldr to newr.
    Pure-constant formulas (=150+136) are untouched."""
    if oldr==newr: return v
    return re.sub(rf'(?<![0-9])([A-Z$]{{1,4}}){oldr}(?![0-9])',
                  lambda m: f'{m.group(1)}{newr}', v)

SRC='Birth_Tracker_Input_EXPANDED.xlsx'
shutil.copy(SRC, SRC+'.bak_expand')          # safety backup

dash=json.load(open('dash_data_unified.json'))['countries']
ES=json.load(open('_eurostat_clean.json'))
EURO_URL='https://ec.europa.eu/eurostat/databrowser/view/demo_fmonth/default/table?lang=en'
MONNAMES=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

# ---- section taxonomy (header text, col-C label) in sheet order ----
SEC=[('EAST ASIA','East Asia'),('SOUTHEAST ASIA','Southeast Asia'),('SOUTH ASIA','South Asia'),
 ('CENTRAL ASIA','Central Asia'),('WESTERN & NORTHERN EUROPE','Western & Northern Europe'),
 ('CENTRAL & EASTERN EUROPE','Central & Eastern Europe'),('NORTH AMERICA','North America'),
 ('LATIN AMERICA & CARIBBEAN','Latin America & Caribbean'),('CAUCASUS & TÜRKIYE','Caucasus & Türkiye'),
 ('MIDDLE EAST & NORTH AFRICA','Middle East & North Africa'),('GULF (NATIONALS)','Gulf (nationals)'),
 ('SUB-SAHARAN AFRICA','Sub-Saharan Africa'),('OCEANIA','Oceania')]
CLABEL={h:c for h,c in SEC}
GULF={'BHR','KWT','OMN','QAT','SAU','ARE'}; CAUC={'ARM','AZE','GEO','TUR'}
BALKAN={'ALB','BIH','HRV','MNE','MKD','SRB','SVN','XKX'}
def section(iso,reg):
    if iso=='CYP': return 'Western & Northern Europe'
    return {
      'Eastern Asia':'East Asia','South-Eastern Asia':'Southeast Asia','Southern Asia':'South Asia',
      'Central Asia':'Central Asia','Northern Africa':'Middle East & North Africa',
      'Sub-Saharan Africa':'Sub-Saharan Africa','Northern America':'North America',
      'Latin America and the Caribbean':'Latin America & Caribbean',
      'Western Europe':'Western & Northern Europe','Northern Europe':'Western & Northern Europe',
      'Eastern Europe':'Central & Eastern Europe','Australia and New Zealand':'Oceania',
      'Melanesia':'Oceania','Micronesia':'Oceania','Polynesia':'Oceania',
    }.get(reg) or (
      ('Gulf (nationals)' if iso in GULF else 'Caucasus & Türkiye' if iso in CAUC else 'Middle East & North Africa')
        if reg=='Western Asia' else
      ('Central & Eastern Europe' if iso in BALKAN else 'Western & Northern Europe')
        if reg=='Southern Europe' else None)

wb=openpyxl.load_workbook(SRC); ws=wb['Births']
wsd=openpyxl.load_workbook(SRC, data_only=True)['Births']   # cached values (for freezing monthly formulas)
MAXC=44

# ---- read existing layout: ordered list of (kind, rowidx_or_data) ----
existing_iso=set(); sec_rows={}; cur=None; order=[]   # order: list of ('hdr',srcrow) or ('data',srcrow)
for r in range(3, ws.max_row+1):
    a=ws.cell(r,1).value; b=ws.cell(r,2).value; c=ws.cell(r,3).value
    if a and not b and not c:            # section header
        cur=a; order.append(('hdr',r)); sec_rows.setdefault(cur,[])
    elif a and b:                        # data row
        existing_iso.add(a); order.append(('data',r)); sec_rows.setdefault(cur,[]).append(r)

# ---- compute new countries per section ----
add={h:[] for h,_ in SEC}
for iso,info in dash.items():
    if iso in existing_iso: continue
    sec=section(iso,info['region'])
    assert sec in CLABEL.values(), (iso,info['region'])
    h=[hh for hh,cc in SEC if cc==sec][0]
    add[h].append((iso,info['name']))
for h in add: add[h].sort(key=lambda x:x[1])

# ---- template styles for brand-new rows: pick an empty European data row (Austria) ----
tmpl_row=None
for r in range(3,ws.max_row+1):
    if ws.cell(r,1).value=='AUT': tmpl_row=r; break
tmpl_style=[copy(ws.cell(tmpl_row,c)._style) for c in range(1,MAXC+1)]

# ---- build new sheet ----
new=wb.create_sheet('Births_new')
def put(nr,c,value,style):
    cell=new.cell(nr,c); cell.value=value
    if style is not None: cell._style=copy(style)
# copy header rows 1 & 2 verbatim
for r in (1,2):
    for c in range(1,MAXC+1):
        sc=ws.cell(r,c); put(r,c,sc.value,sc._style)
nr=3
isorow={}
for h,clabel in SEC:
    # section header: reuse the existing header row's styling for this section
    hdr_src=[rr for k,rr in order if k=='hdr' and ws.cell(rr,1).value==h]
    if hdr_src:
        src=hdr_src[0]
        for c in range(1,MAXC+1):
            sc=ws.cell(src,c); put(nr,c,sc.value,sc._style)
    else:
        put(nr,1,h,ws.cell(tmpl_row,1)._style)
    nr+=1
    # existing data rows (preserve everything; remap self-row formula refs)
    for rr in sec_rows.get(h,[]):
        for c in range(1,MAXC+1):
            sc=ws.cell(rr,c); v=sc.value
            if isinstance(v,str) and v.startswith('='):
                if 4<=c<=39: v=wsd.cell(rr,c).value        # monthly helper-formula -> freeze to cached value (data_only-safe)
                else:        v=remap_formula(v,rr,nr)       # FY auto-sum etc -> keep formula, fix row refs
            put(nr,c,v,sc._style)
        isorow[ws.cell(rr,1).value]=nr
        nr+=1
    # new data rows (match the sheet's FY auto-sum behaviour)
    for iso,name in add[h]:
        for c in range(1,MAXC+1): put(nr,c,None,tmpl_style[c-1])
        put(nr,1,iso,tmpl_style[0]); put(nr,2,name,tmpl_style[1]); put(nr,3,clabel,tmpl_style[2])
        new.cell(nr,41).value=f'=SUM(P{nr}:AA{nr})'
        new.cell(nr,42).value=f'=SUM(AB{nr}:AM{nr})'
        isorow[iso]=nr
        nr+=1
LAST=nr-1

# ---- Eurostat fill (Europe). Only fill empty cells. ----
NUMFMT='#,##0'
filled_log=[]
for iso,e in ES.items():
    if iso not in isorow: continue
    row=isorow[iso]; n24,n25=e['n24'],e['n25']
    if n24==0 and n25==0 and not e['t24'] and not e['t25']: continue
    # 2024 monthly  -> cols 4..15
    for i,v in enumerate(e['m24']):
        if v is None: continue
        cell=new.cell(row,4+i)
        if cell.value in (None,''): cell.value=v; cell.number_format=NUMFMT
    # 2025 monthly  -> cols 16..27
    for i,v in enumerate(e['m25']):
        if v is None: continue
        cell=new.cell(row,16+i)
        if cell.value in (None,''): cell.value=v; cell.number_format=NUMFMT
    # FY2024 total -> AN=40 (manual column, no formula). FY2025 (AO) auto-sums via formula.
    fy24=e['t24'] if e['t24'] is not None else (sum(e['m24']) if n24==12 else None)
    fy25=fy24  # unused placeholder kept for log
    if fy24 is not None and new.cell(row,40).value in (None,''):
        new.cell(row,40).value=fy24; new.cell(row,40).number_format=NUMFMT
    # provenance: only stamp Eurostat where the row has NO source yet (don't override a national source/note)
    if new.cell(row,43).value in (None,''):
        new.cell(row,43).value=EURO_URL
        if new.cell(row,44).value in (None,''):
            if n25==12: y25='2025 full year'
            elif n25>0: y25=f'2025 Jan–{MONNAMES[n25-1]} ({n25} mo)'
            else: y25='2025 not yet posted'
            y24='2024 full year' if (n24==12 or e['t24']) else '2024 n/a'
            new.cell(row,44).value=f'Eurostat demo_fmonth — {y24}; {y25}.'
    filled_log.append((iso,n24,n25,fy24,fy25))

# ---- restore structural formatting ----
for mr in ws.merged_cells.ranges: new.merge_cells(str(mr))
new.freeze_panes=ws.freeze_panes
for c in range(1,MAXC+1):
    L=get_column_letter(c)
    if ws.column_dimensions[L].width: new.column_dimensions[L].width=ws.column_dimensions[L].width
for r in (1,2):
    if ws.row_dimensions[r].height: new.row_dimensions[r].height=ws.row_dimensions[r].height

# ---- swap sheets: remove old Births, rename new, restore position 0 ----
idx=wb.sheetnames.index('Births')
del wb['Births']
new.title='Births'
wb.move_sheet('Births', -(wb.sheetnames.index('Births')-idx))
wb.save(SRC)

# ---- report ----
data_rows=sum(1 for h,_ in SEC for _ in sec_rows.get(h,[])) + sum(len(add[h]) for h,_ in SEC)
print(f"new sheet rows: header2 + {LAST-2} body rows (last row {LAST})")
print(f"data countries: {data_rows} (target 236)")
print(f"added: {sum(len(add[h]) for h,_ in SEC)}")
print(f"Eurostat rows filled: {len(filled_log)}")
print("  sample fills:", [(i,n24,n25) for i,n24,n25,_,_ in filled_log][:6])
