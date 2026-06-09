#!/usr/bin/env python3
"""Tag every country with one of three source categories and attach the right
births estimate. Replaces the earlier flat 'wpp' fill.

  src_cat = 'nso'      -> Verified: national statistical office
                         (reported current-year births, OR a TFR override with a
                          real NSO source)
  src_cat = 'dre'      -> Demoria Research Estimation: figure synthesised from
                         national / official sources (interpolations, worked-around
                         wartime data, etc.). Includes the editor TFR overrides and
                         any country in DRE_FORCE.
  src_cat = 'wpp'      -> UN WPP 2024 only (no recent national override)

Reported-birth countries keep b25/b26/ba and get NO est block.
Everyone else gets est:{prev,cur,year=2025} from DATA's own 2024->2025 series
(override-consistent, so no projection-boundary jumps). Pure-WPP countries' est
is raw WPP; override countries' est is national-data-derived.
"""
import csv,json
from collections import defaultdict
import openpyxl

GLOBE='dhi_globe.html'; DATA_PATH='public/births_data.json'
SHEET='Birth_Tracker_Input_EXPANDED.xlsx'   # DESIGNATION column (AS) is authoritative
# Reported-birth countries that are really Demoria reconstructions, not clean NSO reports
DRE_FORCE={'UKR'}

# ---- read the manual DESIGNATION column (AS = col 45) — the source of truth ----
DESIG_MAP={'NSO':'nso','DRE':'dre','UN WPP 2024':'wpp','UN WPP':'wpp','WPP':'wpp'}
designation={}
try:
    dws=openpyxl.load_workbook(SHEET, data_only=True)['Births']
    for r in range(3, dws.max_row+1):
        iso=dws.cell(r,1).value; lab=dws.cell(r,45).value
        if iso and lab:
            key=str(lab).strip()
            if key in DESIG_MAP: designation[str(iso).strip()]=DESIG_MAP[key]
    print(f"DESIGNATION column read: {len(designation)} countries tagged")
except Exception as e:
    print("WARN could not read DESIGNATION column, falling back to auto-derivation:",e)

# latest override row per country (year, source, url)
ov=defaultdict(list)
with open('overrides/tfr.csv') as f:
    for r in csv.DictReader(f):
        try: yr=int(r['year'])
        except: continue
        ov[r['iso3']].append((yr, (r['source'] or '').strip(), (r['source_url'] or '').strip()))
def latest_override(iso):
    rs=sorted(ov.get(iso,[]))
    return rs[-1] if rs else None
def is_placeholder(src): return src.lower().startswith('wikipedia')

# WPP births from DATA blob
h=open(GLOBE,encoding='utf-8').read()
def blob(name):
    i=h.find(name); s=i+len(name); d=0;p=s
    while p<len(h):
        c=h[p]
        if c=='{': d+=1
        elif c=='}':
            d-=1
            if d==0: return json.loads(h[s:p+1])
        p+=1
DATA=blob('const DATA=')

bd=json.load(open(DATA_PATH,encoding='utf-8'))
tally={'nso':0,'dre':0,'wpp':0}
for c in bd['countries']:
    iso=c['iso']
    c.pop('wpp',None); c.pop('est',None)                 # clear prior fill
    reported = c.get('b26') is not None or c.get('ba')
    lo=latest_override(iso); recent = lo is not None and lo[0]>=2024

    if iso in designation:                 # manual DESIGNATION column wins
        cat=designation[iso]
    elif iso in DRE_FORCE:
        cat='dre'
    elif reported:
        cat='nso'
    elif recent:
        cat='dre' if is_placeholder(lo[1]) else 'nso'
    else:
        cat='wpp'
    c['src_cat']=cat
    # carry the override source name (for display) where it is a real source
    if recent and cat=='nso' and not reported:
        c['tfr_src']=lo[1];
        if lo[2]: c['tfr_src_url']=lo[2]

    # estimate block for non-reporting countries (2024 -> 2025 from DATA)
    if not reported:
        d=DATA.get(iso)
        if d and d.get('ind',{}).get('births') and 2024 in d['yrs'] and 2025 in d['yrs']:
            b24=round(d['ind']['births'][d['yrs'].index(2024)]*1000)
            b25=round(d['ind']['births'][d['yrs'].index(2025)]*1000)
            if b24>0 and b25>0: c['est']={'prev':b24,'cur':b25,'year':2025}
    tally[cat]+=1

bd['src_categories']={'nso':'Verified — national statistical office','dre':'Demoria Research Estimation — synthesised from national & official sources','wpp':'UN WPP 2024 medium-variant projection'}
bd.pop('wpp_note',None)
json.dump(bd,open(DATA_PATH,'w',encoding='utf-8'),ensure_ascii=False,separators=(',',':'))
print("categories:",tally,"total",sum(tally.values()))
# spot checks
by={c['iso']:c for c in bd['countries']}
for iso in ['COL','LKA','NGA','JPN','CHN','TJK','DEU']:
    c=by[iso]; print(f"  {iso}: src_cat={c['src_cat']} est={c.get('est')} reported={c.get('b26') is not None or bool(c.get('ba'))}")
